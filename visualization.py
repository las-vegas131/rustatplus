import io
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, Reference
from collections import defaultdict
import streamlit as st

from config import NEGATIVE_METRICS, METRIC_NAMES_RU, MATCH_METRIC_NAMES_RU
from calculations import get_position_group, percentile_normalize, format_match_metric, build_match_main_table

# -------------------- Радары (Plotly) для интерфейса --------------------
def normalize_for_radar(df, metrics, player_row):
    normed = pd.Series(index=metrics, dtype=float)
    for m in metrics:
        col = df[m]
        mn, mx = col.min(), col.max()
        if mx - mn == 0:
            normed[m] = 0.5
        else:
            val = player_row[m]
            normed[m] = (val - mn) / (mx - mn)
        if m in NEGATIVE_METRICS:
            normed[m] = 1.0 - normed[m]
    return normed

def build_radar_labels(metrics, players, avg_series=None):
    labels = []
    for m in metrics:
        name = METRIC_NAMES_RU.get(m, m) if m.endswith('_p90') or m == 'pass_accuracy' else MATCH_METRIC_NAMES_RU.get(m, m)
        lines = [name]
        for p in players:
            val = p[m]
            if m.endswith('_pct') or m == 'pass_accuracy':
                detail = format_match_metric(m, val, p)
            else:
                detail = format_match_metric(m, val, p)
            lines.append(f"{p['player']}: {detail}")
        if avg_series is not None and m in avg_series:
            avg_val = avg_series[m]
            if m.endswith('_pct') or m == 'pass_accuracy':
                avg_detail = f"{avg_val:.2f}%"
            else:
                avg_detail = f"{avg_val:.2f}"
            lines.append(f"Средние: {avg_detail}")
        labels.append("<br>".join(lines))
    return labels

def add_average_trace(fig, radar_metrics, avg_values, labels, full_df):
    if avg_values is None:
        return
    norm_avg = normalize_for_radar(full_df, radar_metrics, pd.Series(avg_values, index=radar_metrics))
    values = norm_avg.tolist()
    fig.add_trace(go.Scatterpolar(
        r=values + values[:1],
        theta=labels + labels[:1],
        fill='toself',
        name='Средние',
        line=dict(color='#333333', dash='dash'),
        opacity=0.4,
    ))

def create_player_radar_figure(player_row, df, position_weights, avg_values=None):
    pos = get_position_group(player_row['position'])
    weights = position_weights.get(pos, {})
    sorted_metrics = sorted(weights.items(), key=lambda x: -abs(x[1]))
    radar_metrics = [m for m, _ in sorted_metrics if m in df.columns][:8]
    if not radar_metrics:
        radar_metrics = [c for c in df.columns if c.endswith('_p90') or c.endswith('_pct')][:8]
    labels = build_radar_labels(radar_metrics, [player_row], avg_series=avg_values)
    values = normalize_for_radar(df, radar_metrics, player_row).tolist()
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=values + values[:1],
        theta=labels + labels[:1],
        fill='toself',
        name=player_row['player'],
        line_color='blue',
    ))
    if avg_values is not None:
        add_average_trace(fig, radar_metrics, avg_values, labels, df)
    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, 1], showticklabels=False)),
        showlegend=True,
        height=600, width=600,
        margin=dict(l=120, r=120, t=80, b=120),
    )
    return fig

def create_compare_figure(p1, p2, radar_metrics, full_df, avg_values=None):
    players = [p1, p2]
    labels = build_radar_labels(radar_metrics, players, avg_series=avg_values)
    vals1 = normalize_for_radar(full_df, radar_metrics, p1).tolist()
    vals2 = normalize_for_radar(full_df, radar_metrics, p2).tolist()
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=vals1 + vals1[:1],
        theta=labels + labels[:1],
        fill='toself',
        name=f'{p1["player"]} ({p1["rating"]:.1f})',
        line_color='blue', opacity=0.6,
    ))
    fig.add_trace(go.Scatterpolar(
        r=vals2 + vals2[:1],
        theta=labels + labels[:1],
        fill='toself',
        name=f'{p2["player"]} ({p2["rating"]:.1f})',
        line_color='red', opacity=0.6,
    ))
    if avg_values is not None:
        add_average_trace(fig, radar_metrics, avg_values, labels, full_df)
    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, 1], showticklabels=False)),
        showlegend=True,
        height=650, width=650,
        margin=dict(l=120, r=120, t=80, b=120),
        title="Сравнение игроков",
    )
    return fig

def create_position_radar(players_data, full_df, pos_metrics, colors, avg_values=None):
    fig = go.Figure()
    labels = build_radar_labels(pos_metrics[:8], players_data, avg_series=avg_values)
    for i, player_row in enumerate(players_data):
        values = normalize_for_radar(full_df, pos_metrics[:8], player_row).tolist()
        color = colors[i % len(colors)]
        fig.add_trace(go.Scatterpolar(
            r=values + values[:1],
            theta=labels + labels[:1],
            fill='toself',
            name=player_row['player'],
            line_color=color,
            opacity=0.6,
        ))
    if avg_values is not None:
        add_average_trace(fig, pos_metrics[:8], avg_values, labels, full_df)
    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, 1], showticklabels=False)),
        showlegend=True,
        height=900, width=900,
        margin=dict(l=120, r=120, t=80, b=120),
        title="Сравнение по позиции",
    )
    return fig

# -------------------- Экспорт в Excel (формат RuStat) --------------------
def export_matches_advanced(matches_dict, selected_match_ids, team_name, season_year, league_avg, player_season_map, output_bytes_io):
    """
    Экспорт матчей в формат RuStat (один лист, вертикальные заголовки, группировка по игрокам).
    """
    position_ru = {'CB': 'ЦЗ', 'FB': 'КЗ', 'CM': 'НОП', 'CDM': 'НОП',
                   'AM': 'ВОП', 'CAM': 'ВОП', 'CF': 'НАП', 'FW': 'НАП', 'ST': 'НАП'}
    position_order = ['CB', 'FB', 'CM', 'AM', 'FW']

    def safe_int(val):
        if pd.isna(val) or val is None:
            return 0
        try:
            return int(float(val))
        except:
            return 0

    def frac(total, accuracy_pct):
        t = safe_int(total)
        if t == 0:
            return ''
        acc = safe_int(accuracy_pct) if not pd.isna(accuracy_pct) else 0
        succ = int(round(t * acc / 100))
        return f"{t}/{succ}"

    # Сбор данных
    all_rows = []
    for match_id in selected_match_ids:
        data = matches_dict.get(match_id)
        if not data:
            continue
        df = data['df'].copy()
        opponent_label = data.get('label', '')
        if df.empty:
            continue
        df['pos_group'] = df['position'].apply(get_position_group)

        best_worst = {}
        for pos in position_order:
            pos_df = df[df['pos_group'] == pos]
            if pos_df.empty:
                continue
            best_worst[pos] = {}
            metrics_check = ['actions', 'actions_opp_box', 'shots', 'passes', 'progressive_passes',
                             'passes_final_third', 'passes_into_penalty_box', 'crosses', 'challenges',
                             'defensive_challenges', 'attacking_challenges', 'air_challenges', 'tackles', 'dribbles']
            for m in metrics_check:
                if m in pos_df.columns:
                    best_worst[pos][m] = {'max': pos_df[m].max(), 'min': pos_df[m].min()}

        for pos in position_order:
            pos_df = df[df['pos_group'] == pos]
            if pos_df.empty:
                continue
            for _, player_row in pos_df.iterrows():
                player_name = player_row['player']
                minutes = safe_int(player_row.get('minutes', 0))
                season_vals = player_season_map.get(player_name, {}) if player_season_map else {}
                is_best, is_worst = {}, {}
                for m in best_worst.get(pos, {}):
                    val = player_row[m]
                    maxv, minv = best_worst[pos][m]['max'], best_worst[pos][m]['min']
                    is_best[m] = (val == maxv) and (maxv != minv)
                    is_worst[m] = (val == minv) and (maxv != minv)
                row = {
                    'pos': pos, 'player': player_name, 'match': opponent_label, 'minutes': minutes,
                    'ttd_total': safe_int(player_row.get('actions', 0)), 'ttd_succ': safe_int(player_row.get('actions_successful', 0)), 'ttd_raw': player_row.get('actions'),
                    'ttd_opp_total': safe_int(player_row.get('actions_opp_box', 0)), 'ttd_opp_succ': safe_int(player_row.get('actions_opp_box_success', 0)), 'ttd_opp_raw': player_row.get('actions_opp_box'),
                    'shots_target': safe_int(player_row.get('shots_on_target', 0)), 'shots_total': safe_int(player_row.get('shots', 0)),
                    'fouls': safe_int(player_row.get('fouls', 0)), 'fouls_suffered': safe_int(player_row.get('fouls_suffered', 0)),
                    'passes_total': safe_int(player_row.get('passes', 0)), 'passes_acc': player_row.get('pass_accuracy', 0),
                    'prog_total': safe_int(player_row.get('progressive_passes', 0)), 'prog_acc': player_row.get('progressive_passes_accuracy', 0),
                    'ft_total': safe_int(player_row.get('passes_final_third', 0)), 'ft_acc': player_row.get('passes_final_third_accuracy', 0),
                    'pen_total': safe_int(player_row.get('passes_into_penalty_box', 0)), 'pen_acc': player_row.get('passes_into_penalty_box_accuracy', 0),
                    'key_passes': safe_int(player_row.get('key_passes', 0)),
                    'crosses_total': safe_int(player_row.get('crosses', 0)), 'crosses_acc': player_row.get('crosses_accuracy', 0),
                    'chall_total': safe_int(player_row.get('challenges', 0)), 'chall_acc': player_row.get('challenges_won_pct', 0),
                    'def_total': safe_int(player_row.get('defensive_challenges', 0)), 'def_acc': player_row.get('defensive_challenges_won_pct', 0),
                    'att_total': safe_int(player_row.get('attacking_challenges', 0)), 'att_acc': player_row.get('attacking_challenges_won_pct', 0),
                    'air_total': safe_int(player_row.get('air_challenges', 0)), 'air_acc': player_row.get('air_challenges_won_pct', 0),
                    'tack_total': safe_int(player_row.get('tackles', 0)), 'tack_acc': player_row.get('tackles_success_pct', 0),
                    'drib_total': safe_int(player_row.get('dribbles', 0)), 'drib_acc': player_row.get('dribbles_success_pct', 0),
                    'interceptions': safe_int(player_row.get('interceptions', 0)), 'interceptions_opp': safe_int(player_row.get('ball_recoveries_opp_half', 0)),
                    'loose': safe_int(player_row.get('loose_ball_recoveries', 0)), 'loose_opp': safe_int(player_row.get('ball_recoveries_opp_half', 0)),
                    'lost_total': safe_int(player_row.get('lost_balls', 0)), 'lost_own': safe_int(player_row.get('lost_balls_own_half', 0)),
                    'rec_total': safe_int(player_row.get('ball_recoveries', 0)), 'rec_opp': safe_int(player_row.get('ball_recoveries_opp_half', 0)),
                    'is_best': is_best, 'is_worst': is_worst, 'season_vals': season_vals
                }
                all_rows.append(row)

    if not all_rows:
        raise ValueError("Нет данных для экспорта")

    pos_groups = defaultdict(list)
    for r in all_rows:
        pos_groups[r['pos']].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист1'

    default_font = Font(name='Calibri', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = default_font

    ws['A1'] = f"{team_name} {season_year} RuStat Игроки"
    ws['A2'] = "Показатели"
    ws['A3'] = ""
    for cell in ['A1', 'A2']:
        ws[cell].font = Font(name='Calibri', size=11, bold=True)
        ws[cell].alignment = Alignment(horizontal='center')

    headers = [
        'Амплуа', 'Игрок', 'Игра', 'Мин',
        'ТТД/уд', 'ТТД у чужих ворот/уд', 'Удары/в створ',
        'Фолы', 'Фолы на игроке',
        'Передачи/точные', 'Передачи вперёд', 'Передачи в финальную треть',
        'Передачи в финальной трети', 'Передачи в штрафную', 'Передачи ключевые',
        'Навесы', 'Единоборства/удачные', 'В обороне/удачные', 'В атаке/удачные',
        'Вверху/удачные', 'Отборы/удачные', 'Обводки/удачные',
        'Перехваты/на чужой половине', 'Подборы/на чужой половине',
        'Потери/на своей половине', 'Возвраты/на чужо половине'
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=255, wrapText=False)
        if col_idx <= 3:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 6
    ws.row_dimensions[4].height = 120
    ws.freeze_panes = 'B5'

    row_idx = 5
    for pos in position_order:
        rows_list = pos_groups.get(pos, [])
        if not rows_list:
            continue
        player_matches = defaultdict(list)
        for r in rows_list:
            player_matches[r['player']].append(r)
        for player in sorted(player_matches.keys()):
            matches = sorted(player_matches[player], key=lambda x: x['match'])
            first_match = True
            for match in matches:
                ws.cell(row=row_idx, column=1, value=position_ru.get(pos, pos))
                ws.cell(row=row_idx, column=2, value=player if first_match else '')
                ws.cell(row=row_idx, column=3, value=match['match'])
                ws.cell(row=row_idx, column=4, value=match['minutes'])

                def format_cell(value, raw_val, avg_val, metric_name, is_best_flag, is_worst_flag, season_val):
                    base = str(value) if value else ''
                    if is_best_flag:
                        base = f"🟢 {base}"
                    elif is_worst_flag:
                        base = f"🔴 {base}"
                    if season_val is not None and not pd.isna(season_val):
                        try:
                            s_val = float(season_val)
                            if metric_name in NEGATIVE_METRICS:
                                if raw_val < s_val:
                                    base += " ↑"
                                elif raw_val > s_val:
                                    base += " ↓"
                            else:
                                if raw_val > s_val:
                                    base += " ↑"
                                elif raw_val < s_val:
                                    base += " ↓"
                        except:
                            pass
                    marker = ''
                    color = None
                    if raw_val is not None and avg_val is not None:
                        try:
                            rf = float(raw_val)
                            af = float(avg_val)
                            if metric_name in NEGATIVE_METRICS:
                                if rf < af:
                                    marker, color = "🔵 ", "0000FF"
                                elif rf > af:
                                    marker, color = "🟤 ", "8B4513"
                            else:
                                if rf > af:
                                    marker, color = "🔵 ", "0000FF"
                                elif rf < af:
                                    marker, color = "🟤 ", "8B4513"
                        except:
                            pass
                    return (marker + base if marker else base), color

                ttd_val = f"{match['ttd_total']}/{match['ttd_succ']}" if match['ttd_total'] > 0 else ''
                final, col = format_cell(ttd_val, match['ttd_raw'], league_avg.get('actions') if league_avg else None,
                                         'actions', match['is_best'].get('actions', False), match['is_worst'].get('actions', False),
                                         match['season_vals'].get('actions') if 'actions' in match['season_vals'] else None)
                cell = ws.cell(row=row_idx, column=5, value=final)
                if col:
                    cell.font = Font(color=col, name='Calibri', size=11)

                ttd_opp_val = f"{match['ttd_opp_total']}/{match['ttd_opp_succ']}" if match['ttd_opp_total'] > 0 else ''
                final, col = format_cell(ttd_opp_val, match['ttd_opp_raw'], league_avg.get('actions_opp_box') if league_avg else None,
                                         'actions_opp_box', match['is_best'].get('actions_opp_box', False), match['is_worst'].get('actions_opp_box', False),
                                         match['season_vals'].get('actions_opp_box') if 'actions_opp_box' in match['season_vals'] else None)
                cell = ws.cell(row=row_idx, column=6, value=final)
                if col:
                    cell.font = Font(color=col, name='Calibri', size=11)

                ws.cell(row=row_idx, column=7, value=f"{match['shots_target']}/{match['shots_total']}" if match['shots_total'] > 0 else '')
                ws.cell(row=row_idx, column=8, value=match['fouls'] if match['fouls'] > 0 else '')
                ws.cell(row=row_idx, column=9, value=match['fouls_suffered'] if match['fouls_suffered'] > 0 else '')
                ws.cell(row=row_idx, column=10, value=frac(match['passes_total'], match['passes_acc']))
                ws.cell(row=row_idx, column=11, value=frac(match['prog_total'], match['prog_acc']))
                ws.cell(row=row_idx, column=12, value=frac(match['ft_total'], match['ft_acc']))
                ws.cell(row=row_idx, column=13, value=frac(match['ft_total'], match['ft_acc']))
                ws.cell(row=row_idx, column=14, value=frac(match['pen_total'], match['pen_acc']))
                ws.cell(row=row_idx, column=15, value=match['key_passes'] if match['key_passes'] > 0 else '')
                ws.cell(row=row_idx, column=16, value=frac(match['crosses_total'], match['crosses_acc']))
                ws.cell(row=row_idx, column=17, value=frac(match['chall_total'], match['chall_acc']))
                ws.cell(row=row_idx, column=18, value=frac(match['def_total'], match['def_acc']))
                ws.cell(row=row_idx, column=19, value=frac(match['att_total'], match['att_acc']))
                ws.cell(row=row_idx, column=20, value=frac(match['air_total'], match['air_acc']))
                ws.cell(row=row_idx, column=21, value=frac(match['tack_total'], match['tack_acc']))
                ws.cell(row=row_idx, column=22, value=frac(match['drib_total'], match['drib_acc']))
                ws.cell(row=row_idx, column=23, value=f"{match['interceptions']}/{match['interceptions_opp']}" if match['interceptions'] > 0 else '')
                ws.cell(row=row_idx, column=24, value=f"{match['loose']}/{match['loose_opp']}" if match['loose'] > 0 else '')
                ws.cell(row=row_idx, column=25, value=f"{match['lost_total']}/{match['lost_own']}" if match['lost_total'] > 0 else '')
                ws.cell(row=row_idx, column=26, value=f"{match['rec_total']}/{match['rec_opp']}" if match['rec_total'] > 0 else '')

                row_idx += 1
                first_match = False

    for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=26):
        for cell in row:
            if cell.column in [1, 2, 3]:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_bytes_io)
    st.success("Экспорт RuStat завершён")

# -------------------- Стандартный экспорт матча с нативными радарами Excel --------------------
def export_match_standard_with_charts(df_match, selected_metrics, position_tables, league_avg, output_bytes_io):
    """
    Экспорт матча в Excel: лист 'Данные' + лист 'Радары' с диаграммами (RadarChart через openpyxl).
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'Данные'

    # 1. Общая таблица
    df_main = build_match_main_table(df_match, selected_metrics,
                                     league_avg=league_avg,
                                     player_season_map=None)
    for r_idx, row in enumerate(df_main.values, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=value)
    for c_idx, header in enumerate(df_main.columns, start=1):
        cell = ws_data.cell(row=1, column=c_idx, value=header)
        cell.font = Font(bold=True)

    # 2. Таблицы по позициям
    next_row = len(df_main) + 3
    for pos, (rows, headers) in position_tables.items():
        if rows:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_data.cell(row=next_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            next_row += 1
            for row in rows:
                for col_idx, val in enumerate(row, start=1):
                    ws_data.cell(row=next_row, column=col_idx, value=val)
                next_row += 1
            next_row += 1

    # 3. Лист с радарами (Excel native)
    ws_radar = wb.create_sheet("Радары")
    ws_radar.column_dimensions['A'].width = 30
    current_row = 1

    # Определяем метрики для радара (все числовые, кроме ТТД и служебных)
    exclude = ['player', 'position', 'minutes', 'team', 'league', 'pos_group', 'ttd_actions', 'ttd_opp_actions']
    all_numeric = [m for m in df_match.columns if m not in exclude and pd.api.types.is_numeric_dtype(df_match[m])]
    radar_metrics = all_numeric[:12]  # не более 12 метрик для читаемости
    if not radar_metrics:
        radar_metrics = [m for m in df_match.columns if m.endswith('_p90')][:8]

    # Для каждого игрока создаём радар
    for _, player_row in df_match.iterrows():
        player_name = player_row['player']
        # Вычисляем нормализованные значения (0..1) для метрик
        norm_values = []
        avg_values = []
        for m in radar_metrics:
            if m in df_match.columns:
                min_val = df_match[m].min()
                max_val = df_match[m].max()
                if max_val - min_val == 0:
                    norm = 0.5
                else:
                    norm = (player_row[m] - min_val) / (max_val - min_val)
                    if m in NEGATIVE_METRICS:
                        norm = 1.0 - norm
                norm_values.append(norm)
                # Среднее по лиге (если есть)
                if league_avg and m in league_avg:
                    avg = league_avg[m]
                    if max_val - min_val == 0:
                        avg_norm = 0.5
                    else:
                        avg_norm = (avg - min_val) / (max_val - min_val)
                        if m in NEGATIVE_METRICS:
                            avg_norm = 1.0 - avg_norm
                    avg_values.append(avg_norm)
                else:
                    avg_values.append(0.5)
            else:
                norm_values.append(0.5)
                avg_values.append(0.5)

        # Записываем данные на лист
        start_row = current_row
        ws_radar.cell(row=start_row, column=1, value=player_name)
        ws_radar.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        start_row += 1
        # Заголовки метрик
        for i, m in enumerate(radar_metrics, start=1):
            ws_radar.cell(row=start_row, column=i, value=MATCH_METRIC_NAMES_RU.get(m, m))
        start_row += 1
        # Значения игрока
        for i, val in enumerate(norm_values, start=1):
            ws_radar.cell(row=start_row, column=i, value=val)
        start_row += 1
        # Средние значения
        for i, val in enumerate(avg_values, start=1):
            ws_radar.cell(row=start_row, column=i, value=val)
        start_row += 1

        # Создаём радарную диаграмму
        radar = RadarChart()
        radar.add_data(Reference(ws_radar, min_row=start_row-2, max_row=start_row-1, min_col=1, max_col=len(radar_metrics)), titles_from_data=False)
        radar.set_categories(Reference(ws_radar, min_row=start_row-3, max_row=start_row-3, min_col=1, max_col=len(radar_metrics)))
        radar.title = player_name
        radar.style = 26
        radar.y_axis.title = "Нормализованное значение"
        radar.x_axis.title = "Метрики"
        # Размещаем диаграмму на листе
        pos_row = start_row + 1
        ws_radar.add_chart(radar, f"A{pos_row}")
        # Увеличиваем высоту строк для диаграммы
        for i in range(20):
            ws_radar.row_dimensions[pos_row + i].height = 15
        current_row = pos_row + 25

    # Автоширина для листа 'Данные'
    for col in ws_data.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
        ws_data.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_bytes_io)
    st.success("Стандартный экспорт с радарами (Excel native) завершён")
